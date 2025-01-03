from collections import defaultdict
from concurrent.futures import ProcessPoolExecutor, as_completed
from concurrent.futures.thread import ThreadPoolExecutor
from typing import List, Dict

import pandas
from PySide6 import QtWidgets, QtCore
import pandas as pd
import openpyxl
import os
from datetime import datetime

from PySide6.QtCore import QThread, Signal, Qt
from PySide6.QtGui import QBrush, QColor
from PySide6.QtWidgets import QTableWidgetItem, QLabel, QSpinBox, QFileDialog
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from matplotlib.figure import Figure
from matplotlib.backends.backend_qt5agg import (
    FigureCanvasQTAgg as FigureCanvas,
    NavigationToolbar2QT as NavigationToolbar,
)


class UiMixin(object):
    def toggle_selection(self, item):
        if item.checkState() == QtCore.Qt.Checked:
            item.setCheckState(QtCore.Qt.Unchecked)
        else:
            item.setCheckState(QtCore.Qt.Checked)

    # 单选模式
    def single_selection(self, widget, item):
        if item.checkState() == QtCore.Qt.Checked:
            item.setCheckState(QtCore.Qt.Unchecked)
        else:
            # 取消其他选中项
            for index in range(widget.count()):
                it = widget.item(index)
                if it.checkState() == QtCore.Qt.Checked:
                    it.setCheckState(QtCore.Qt.Unchecked)
            item.setCheckState(QtCore.Qt.Checked)

    def previous_tab(self, tab):
        current_index = tab.currentIndex()
        if current_index > 0:
            tab.setCurrentIndex(current_index - 1)

    def next_tab(self, tab):
        current_index = tab.currentIndex()
        if current_index < tab.count() - 1:
            tab.setCurrentIndex(current_index + 1)

    def select_all_files(self, list_widget):
        for index in range(list_widget.count()):
            item = list_widget.item(index)
            item.setCheckState(QtCore.Qt.Checked)

    def deselect_all_files(self, list_widget):
        for index in range(list_widget.count()):
            item = list_widget.item(index)
            item.setCheckState(QtCore.Qt.Unchecked)


class GenericWorkerThread(QThread):
    progress = Signal(int)  # 进度更新信号，发送一个整数
    resultReady = Signal(object)  # 计算结果信号，可以发送任何对象
    errorOccurred = Signal(str)  # 错误信号，发送错误信息

    def __init__(self, task_func, *args, **kwargs):
        super().__init__()
        self.task_func = task_func
        self.args = args
        self.kwargs = kwargs

    def run(self):
        try:
            result = self.task_func(*self.args, **self.kwargs)
            self.resultReady.emit(result)
        except Exception as e:
            self.errorOccurred.emit(str(e))


class FileTools(object):
    def extract_from_filename(self, filename):
        try:
            parts = filename.split('_')
            client_name = parts[0]
            date_str = parts[2]
            date_obj = datetime.strptime(date_str, '%m.%d.%Y')
            return client_name, date_obj.strftime('%m-%d-%Y')
        except (IndexError, ValueError):
            return None, None

    def load_files_by_folder(self, folder_path: str):
        date_to_files_map = defaultdict(list)
        client_name = None
        for file in os.listdir(folder_path):
            if file.endswith(".csv") or file.endswith(".xlsx"):
                new_client_name, date_str = self.extract_from_filename(file)
                client_name = client_name or new_client_name
                if not date_str:
                    continue
                date_to_files_map[date_str].append(os.path.join(folder_path, file))
        return client_name, date_to_files_map

    def load_files_data_frames_map(
            self,
            file_paths_map,
            select_dates: List[str],
            select_fields: List[str] = None
    ):
        dfs_map = defaultdict(list)
        read_params = {}
        if select_fields:
            read_params['usecols'] = select_fields
        else:
            read_params['nrows'] = 0

        # 定义一个读取文件的辅助函数
        def read_file(file_path, date_str):
            try:
                if file_path.endswith(".csv"):
                    df = pd.read_csv(file_path, **read_params)
                elif file_path.endswith(".xlsx"):
                    sheets = pd.read_excel(file_path, sheet_name=None, engine='openpyxl', **read_params)
                    df = pd.concat(sheets.values(), ignore_index=True)  # 将所有工作表合并为一个 DataFrame
                else:
                    raise ValueError("Unsupported file format.")
                return date_str, df
            except Exception as e:
                raise Exception(f"Can't read file: {file_path} by error: {e}")

        with ThreadPoolExecutor(os.cpu_count() * 2) as executor:
            future_to_date = {}
            for date_str in select_dates:
                file_paths = file_paths_map.get(date_str, [])
                for file_path in file_paths:
                    future = executor.submit(read_file, file_path, date_str)
                    future_to_date[future] = date_str

            # 收集结果并存入 dfs_map
            for future in as_completed(future_to_date):
                try:
                    date_str, df = future.result()
                    dfs_map[date_str].append(df)
                except Exception as e:
                    print(e)

        return dfs_map

    def load_data_frames_headers(self, data_frames):
        headers_set = []
        for df_list in data_frames:
            for df in df_list:
                for col in df.columns:
                    if col not in headers_set:
                        headers_set.append(col)
        return headers_set


class MathTools(object):
    def uniqueness_test(self, data_frames_map: Dict[str, List[pd.DataFrame]], select_dates: List[str], select_fields: List[str],
                        audit_field: str = None, top_exceptions: int = 10):
        uniqueness_results, uniqueness_exceptions = [], []
        audit_field_sum_cache = {}
        except_count = 0
        for date_str in select_dates:
            dfs = data_frames_map[date_str]
            new_df = pd.concat(dfs, ignore_index=True)
            selected_df = new_df[select_fields]
            # 如果audit_field提供并存在，将其总和计算并缓存
            if audit_field and audit_field in new_df.columns:
                if date_str not in audit_field_sum_cache:
                    audit_field_sum_cache[date_str] = new_df[audit_field].sum()
                audit_field_sum = audit_field_sum_cache[date_str]
            else:
                audit_field_sum = None

            # 如果有多个字段组合，创建组合列（只计算一次）
            if len(select_fields) > 1:
                new_col_name = "_EY_".join(select_fields)
                selected_df = selected_df.assign(
                    **{new_col_name: selected_df[select_fields].apply(lambda x: "_".join(x.astype(str)), axis=1)})

            # 对所有列进行唯一性测试
            for col in selected_df.columns:
                unique_count = selected_df[col].nunique()
                total_count = len(selected_df[col])
                single_period_result = "Unique" if unique_count == total_count else "Not Unique"
                uniqueness_results.append((date_str, col, None, single_period_result))

                # 对于非唯一的情况，记录每行的 audit_field 的值及其统计信息
                if unique_count == total_count:
                    continue
                duplicate_mask = selected_df.duplicated(subset=[col], keep=False)
                duplicates = selected_df[duplicate_mask]
                duplicates[col] = duplicates[col].fillna('NaN')

                if except_count >= top_exceptions:
                    continue
                for dup_value, group in duplicates.groupby(col):
                    dup_count = group.shape[0]
                    if except_count >= top_exceptions:
                        break
                    for idx, row in group.iterrows():
                        field_value = dup_value if dup_value != 'NaN' else None
                        audit_value = new_df.at[idx, audit_field] if audit_field and idx in new_df.index else None
                        try:
                            ratio = round(audit_value / audit_field_sum, 4) if audit_value is not None and audit_field_sum else 0
                        except ZeroDivisionError:
                            ratio = 0
                        # Append exception details    
                        uniqueness_exceptions.append(
                            (date_str, col, field_value, dup_count, audit_value, audit_field_sum, ratio))
                        except_count += 1
                        if except_count >= top_exceptions:
                            break
        uniqueness_results_sorted = sorted(uniqueness_results, key=lambda x: (x[1], x[0]))

        from collections import defaultdict

        col_results = defaultdict(list)
        for result in uniqueness_results_sorted:
            col_results[result[1]].append(result)

        final_results = []
        for col, results in col_results.items():
            all_unique = all(result[3] == "Unique" for result in results)
            overall_result = "Always unique" if all_unique else "Not Always Unique"
            for result in results:
                final_result = (result[0], result[1], overall_result, result[3])
                final_results.append(final_result)

        return final_results, uniqueness_exceptions

    # def omitted_records_test(self, data_frames_map: Dict[str, List[pd.DataFrame]], select_dates: List[str], select_fields: List[str]):
  
    # # Ensure at least three dates are selected for meaningful analysis
    #  if len(select_dates) < 3:
    #     return {}, []

    #  omitted_records_summary = defaultdict(int)
    #  omitted_records_details = []

    #  try:
    #     # Sort dates and create a mapping to indices
    #     sorted_dates = sorted(select_dates, key=lambda x: datetime.strptime(x, '%m-%d-%Y'))
    #     date_index = {date: index for index, date in enumerate(sorted_dates)}

    #     for field in select_fields:
    #         data_map = defaultdict(list)

    #         # Collect values for each field across dates
    #         for date_str in sorted_dates:
    #             dfs = data_frames_map.get(date_str, [])
    #             for df in dfs:
    #                 if field not in df.columns:
    #                     print(f"Warning: Field '{field}' missing in data for date {date_str}. Skipping...")
    #                     continue

    #                 values = df[field].dropna()  # Drop NaN values
    #                 for value in values:
    #                     data_map[value].append(date_str)

    #         for value, dates in data_map.items():
    #             # Skip values present in fewer than 2 periods or in all periods
    #             if len(dates) < 2 or len(dates) == len(sorted_dates):
    #                 continue

    #             # Find missing intermediate dates
    #             date_set = set(dates)
    #             start_date, end_date = dates[0], dates[-1]
    #             start_idx = date_index[start_date]
    #             end_idx = date_index[end_date]
    #             split_dates = sorted_dates[start_idx:end_idx + 1]
    #             diff_dates = [date for date in split_dates if date not in date_set]

    #             if diff_dates:
    #                 omitted_records_details.append([value, len(diff_dates), ', '.join(diff_dates)])
    #                 for date in diff_dates:
    #                     omitted_records_summary[date] += 1

    #  except Exception as e:
    #     print(f"Error during omitted records test: {e}")
    #     raise

    #  return dict(omitted_records_summary), omitted_records_details
    def omitted_records_test(self, data_frames_map: Dict[str, List[pd.DataFrame]], select_dates: List[str], select_fields: List[str]):
    # Ensure at least three dates are selected for meaningful analysis
     if len(select_dates) < 3:
        return {}, []

     omitted_records_summary = defaultdict(int)
     omitted_records_details = []

     try:
        # Sort dates and create a mapping to indices
        sorted_dates = sorted(select_dates, key=lambda x: datetime.strptime(x, '%m-%d-%Y'))
        date_index = {date: index for index, date in enumerate(sorted_dates)}

        for field in select_fields:
            data_map = defaultdict(list)

            # Collect unique field values for each date
            for date_str in sorted_dates:
                dfs = data_frames_map.get(date_str, [])
                for df in dfs:
                    if field not in df.columns:
                        print(f"Warning: Field '{field}' missing in data for date {date_str}. Skipping...")
                        continue

                    # Collect unique non-null values
                    values = df[field].dropna().unique()
                    for value in values:
                        data_map[value].append(date_str)

            # Analyze missing intermediate dates for each field value
            for value, dates in data_map.items():
                # Skip values that appear in fewer than 2 periods or all periods
                if len(dates) < 2 or len(dates) == len(sorted_dates):
                    continue

                # Find missing intermediate dates
                date_set = set(dates)
                first_date, last_date = dates[0], dates[-1]
                intermediate_dates = sorted_dates[date_index[first_date]: date_index[last_date] + 1]
                missing_dates = [date for date in intermediate_dates if date not in date_set]

                if missing_dates:
                    omitted_records_details.append([value, len(missing_dates), ', '.join(missing_dates)])
                    for date in missing_dates:
                        omitted_records_summary[date] += 1

     except Exception as e:
        print(f"Error during omitted records test: {e}")
        raise

     return dict(omitted_records_summary), omitted_records_details


    def omitted_records_analysis(self):
        pass

    def rollforward_analysis(self,data_frames_map, select_dates, unique_id_field, audit_field):
 
     rollforward_summary = []
     sorted_dates = sorted(select_dates, key=lambda x: datetime.strptime(x, '%m-%d-%Y'))

     for i in range(len(sorted_dates) - 1):
        prev_date, curr_date = sorted_dates[i], sorted_dates[i + 1]
        prev_df = pd.concat(data_frames_map[prev_date], ignore_index=True)
        curr_df = pd.concat(data_frames_map[curr_date], ignore_index=True)

        prev_ids = set(prev_df[unique_id_field])
        curr_ids = set(curr_df[unique_id_field])

        added_ids = curr_ids - prev_ids
        dropped_ids = prev_ids - curr_ids
        remained_ids = prev_ids & curr_ids

        added_balance = curr_df[curr_df[unique_id_field].isin(added_ids)][audit_field].sum()
        dropped_balance = prev_df[prev_df[unique_id_field].isin(dropped_ids)][audit_field].sum()
        rolling_start_balance = prev_df[prev_df[unique_id_field].isin(remained_ids)][audit_field].sum()
        rolling_end_balance = curr_df[curr_df[unique_id_field].isin(remained_ids)][audit_field].sum()

        rollforward_summary.append({
            "Previous Date": prev_date,
            "Current Date": curr_date,
            "Added Count": len(added_ids),
            "Added Balance": added_balance,
            "Dropped Count": len(dropped_ids),
            "Dropped Balance": dropped_balance,
            "Remained Count": len(remained_ids),
            "Rolling Start Balance": rolling_start_balance,
            "Rolling End Balance": rolling_end_balance,
            "Rolling Balance Change": rolling_end_balance - rolling_start_balance
        })

     return rollforward_summary


    def profiling_analysis(self, data_frames_map, select_dates, select_fields):
        """
        Perform profiling analysis on the selected fields.

        Parameters:
            data_frames_map (dict): Mapping of dates to dataframes.
            select_dates (list): List of dates for profiling.
            select_fields (list): List of fields to profile.

        Returns:
            dict: A dictionary containing profiling results by field.
        """
        profiling_results = {}

        for date in select_dates:
            profiling_results[date] = []
            df = pd.concat(data_frames_map[date], ignore_index=True)

            for field in select_fields:
                if field not in df.columns:
                    continue
                
                field_stats = {"Field Name": field}
                column = df[field]

                # Basic statistics
                field_stats["Count"] = len(column)
                field_stats["Count Null"] = column.isnull().sum()
                field_stats["% Null"] = round(column.isnull().mean() * 100, 2)
                field_stats["Count Distinct"] = column.nunique()
                field_stats["Count Unique"] = (column.value_counts() == 1).sum()

                if pd.api.types.is_numeric_dtype(column):
                    field_stats.update({
                        "Minimum": column.min(),
                        "Maximum": column.max(),
                        "Sum": column.sum(),
                        "Simple Average": column.mean(),
                        "Count Negative": (column < 0).sum(),
                        "% Negative": round((column < 0).mean() * 100, 2),
                        "Count Zero": (column == 0).sum(),
                        "% Zero": round((column == 0).mean() * 100, 2),
                    })

                elif pd.api.types.is_datetime64_any_dtype(column):
                    field_stats.update({
                        "Minimum": column.min(),
                        "Maximum": column.max(),
                        "Range": (column.max() - column.min()).days if column.max() and column.min() else None,
                    })

                elif pd.api.types.is_string_dtype(column):
                    mode = column.mode()
                    field_stats.update({
                        "Mode": mode.iloc[0] if not mode.empty else None
                    })

                profiling_results[date].append(field_stats)

        return profiling_results


class Constants:
    UNIQUENESS_SUMMARY_HEADERS = ["EY SnapDate", "Tested Field", "All Period Results", "Single Period Results"]
    UNIQUENESS_EXCEPTIONS_HEADERS = ["EY SnapDate", "Tested Field", "Field Value", "Count Value", "Audit Balance",
                                     "Total Audit Balance", "% of Total Audit Balance"]
    OMITTED_RECORDS_SUMMARY_HEADERS = ["EY SnapDate", "Count Omitted Records"]
    OMITTED_RECORDS_DETAILS_HEADERS = ["Unique ID", "Count Skipped EY SnapDates", "Skipped EY SnapDates"]

    Instructions = {
        "Data Pre Processing": """
<b>Data Terminology</b><br>
1. <u>Data Record:</u> A single item/unit in the data set (i.e., row) characterized by all the values observed in each data field (e.g., Account #).<br>
2. <u>Data Field:</u> The nature of the dimension or measure (e.g., Account Type, Interest Rate); typically that described by the field header, it is that reflected in the values of a column.<br>
3. <u>Field Value:</u> Irrespective of data type, the precise value observed in a given data field for a given data record (e.g., Commercial; 3/31/202X; $100.04). Terms "null," "empty," "blank" refer to absence of a field value in the data field of a record (and thus differ from $0.00).<br>
4. <u>Field Type:</u> The format in which the field value is stored within the client's data (e.g., Text ("String"), Date, Number).<br>
5. <u>kipped/Omitted records:</u> Records (Unique IDs) appearing in non-successive periods but missing in at least one intermediate period are said to be omitted or skipped. For instance, if a record appears in Periods 1 and 3 but not in Period 2, it is considered an omitted record. However, if a record appears in successive periods, such as Periods 2 and 3, but is missing in Periods 1 or 5, it is not flagged as an omitted record because the record is not absent in any intermediate periods.<br><br>

<b>Analysis Terminology</b><br>   
1. <u>Uniqueness Analysis:</u> Evaluate records to ensure they are distinct and free from duplicates.<br>
2. <u>Omitted Records Analysis:</u>  Identify and analyze any missing or skipped records.<br>
3. <u>Cross-Period Analysis:</u> Examine data trends and changes across multiple reporting periods to ensure consistency and accuracy.<br>
4. <u>Profile Analysis:</u> Analyze and interpret key characteristics and attributes of the dataset to gain meaningful insights.<br><br>
Click "<i>Next</i>" to proceed once you have completed the analysis or "<i>Cancel</i>" to exit the process. 

""",
        "Read Folder": """
1. Click "<i>Select Folder</i>" to choose a directory containing <i>.csv</i> or <i>.xslx</i> files for data validation.<br>
2. Once the folder is selected, the application will display the dats (<i>MM.DD.YYYY</i>) from the file names. Select the files you wish to include in the data validation process by marking the corresponding boxes.<br>
    &nbsp;&nbsp;&nbsp;&nbsp;2.1 Use "<i>Select All</i>" to include all files.<br>
    &nbsp;&nbsp;&nbsp;&nbsp;2.2 Use "<i>Deselect All</i>" to clear all selections.<br>
3. Click "<i>Next</i>" or navigate through the tabs at the top to perform additional analyses or "<i>Cancel</i>" to exit the process..
       """,
        "Uniqueness Analysis": """
1. Select the fields from the left list box that you want to test for uniqueness. These fields will be checked to ensure there are no duplicate values.<br>
    &nbsp;&nbsp;&nbsp;&nbsp;1.1 Designate the field used by the client to uniquely identify each individual record at and across all reporting periods. <br>
    &nbsp;&nbsp;&nbsp;&nbsp;1.2 On the rare occasion that the client uses the combination of values in two or more fields as the unique identifier, select each such field to test the EY concatenated value for uniqueness (field value concatenation should not be employed unless confirmed by client).<br>
2. Select the field from the right list box that represents the balance or activity to be audited. This field will be used in the App for calculating metrics like weighted averages, percentages, and totals.<br>
3. Use the "<i>Top Exceptions</i>" texbox to specify how many exceptions you want to display in the results. The maximum allowable limit is 20,000.<br>
3. Click "<i>Perform Uniqueness Test</i>" to analyze the selected fields. The results will be shown in the tabs below:<br>
    &nbsp;&nbsp;&nbsp;&nbsp;3.1 Uniqueness Field Summary: A summary of the uniqueness check across the selected fields.<br>
    &nbsp;&nbsp;&nbsp;&nbsp;3.2 Uniqueness Exceptions: Detailed information about any exceptions found during the analysis.<br>
4. If you wish to save the test results, click "<i>Export Uniqueness Test</i>".<br>
5. Click "<i>Next</i>" to proceed once you have completed the analysis or "<i>Cancel</i>" to exit the process.<br><br>

<font color="red"><b>Attention:</b> Though results are produced for each selected field, the concatenated value is treated as the unique identifier for all other analyses. As with no tested fields are unique, results indicating multiple fields are unique also warrant clarification with the client as to the true unique identifier.</font>

""",
        "Omitted Records Analysis": """
1. Click the "<i>Perform Omitted Records Test</i>" to initiate the analysis. This test identifies any records that appear in non-successive periods but are missing in at least one intermediate period.<br>
2. Once the analysis is complete, you can view the results in the following tabs:<br>
    &nbsp;&nbsp;&nbsp;&nbsp;2.1 Omitted Records Summary: Provides a summary of the records identified as omitted.<br>
    &nbsp;&nbsp;&nbsp;&nbsp;2.2 Omitted Records Details: Displays detailed information about each omitted record, including specifics about the periods in which the records were absent.<br>
3. To save the analysis results, click the "<i>Export Omitted Records Test</i>" button. This will export both the summary and detailed results for further review.<br>
4. Click "<i>Next</i>" to proceed once you have completed the analysis or "<i>Cancel</i>" to exit the process.<br><br>
<font color="red"><b>Attention:</b> Unnecessary concatenation of field values may yield "false-positives" in the omitted records analysis, due to the Unique ID's critical role in linking each record across periods. If there is no proper Unique ID, the omitted records analysis will produce inaccurate results, potentially flagging records incorrectly as omitted. It is crucial that we investigate and resolve the cause of omitted records, including cases arising from inappropriate use of concatenation (i.e., not following the client's intended approach).</font>


""",
        "Cross-Period Analysis": """
1. Click the "<i>Generate a Rollforward</i>" button to create a summary that identifies reconciling items and highlights changes in records within the specified snapshot date population based on the designated Unique ID.<br>
2. Review changes in record counts and balances to evaluate whether any discrepancies, such as unexpected record "churn" or "turnover," deviate from expectations or indicate a lack of comparability.<br>
3. To save the analysis results, click the "<i>Export Rollfoward Summary</i>" button.<br>
4. Click "<i>Next</i>" to proceed once you have completed the analysis or "<i>Cancel</i>" to exit the process.<br><br>

<font color="red"><b>Attention:</b> Unnecessary concatenation of field values may impair the efficacy of the rollfoward analysis because of the Unique ID's critical role linking each record across periods. It is crucial that users avoid inappropriate use of concatenation (i.e., use other than when employed by the client), so as to ensure the field values for all records are included in evaluation.</font>
""",
        "Profile Analysis": """

To complete prescribed field profiling and validations for the specified field type, users select:<br>
1. Select Date field/s to be analyzed <br>
    &nbsp;&nbsp;&nbsp;&nbsp;1.1 Client data fields representing date values (e.g., origination date). If present, the data effective date field ("as-of") should always be selected (if not present, confirm appropriate to append based on report date and instruct DID accordingly).<br>
2. Select Number field/s to be analyzed <br>
    &nbsp;&nbsp;&nbsp;&nbsp;2.1 Client fields representing numerical values such as balances, transaction amounts, rates, or other figures used in mathematical functions (e.g., addition, multiplication). Balances, rates, etc., key to our audit procedures, typically warrant selection; however, fields comprised of numbers used as identifiers should be designated "Text" (e.g., account ID - 1234567).<br>
3. Select Text (String) field/s to be analyzed <br>
    &nbsp;&nbsp;&nbsp;&nbsp;3.1 Client data fields representing categories, names, places, and strings of alphanumeric characters, including where values are comprised of numbers that do not represent units of measure (e.g., account numbers). <br><br>

<font color="red"><b>Attention:</b> Consider the relevance of each field to our audit procedures; select fields from each of the three field types first before proceeding to mitigate the risk of incomplete or improper designation of a field, version control issues, file overwrites, etc.</font>
"""
    }


class FileConsolidationApp(QtWidgets.QWidget, UiMixin):
    def __init__(self):
        super().__init__()
        self.file_tools = FileTools()
        self.math_tools = MathTools()
        self.setWindowTitle("Data Validation Tool")
        self.resize(900, 700)

        self.main_layout = QtWidgets.QVBoxLayout(self)
        self.instruction_frame = QtWidgets.QGroupBox("Instructions")
        self.instruction_layout = QtWidgets.QVBoxLayout()
        self.instruction_frame.setLayout(self.instruction_layout)

        self.instruction_label = QtWidgets.QLabel(list(Constants.Instructions.values())[0])
        self.instruction_label.setWordWrap(True)
        self.instruction_layout.addWidget(self.instruction_label)
        self.main_layout.addWidget(self.instruction_frame)

        # Tab Widget for Analysis
        self.tab_widget = QtWidgets.QTabWidget()
        self.main_layout.addWidget(self.tab_widget)

        self.create_tabs()
        self.button_layout = QtWidgets.QHBoxLayout()
        self.button_layout.addStretch()  # Add stretch to push buttons to the right
        self.back_button = QtWidgets.QPushButton("< Back")
        self.next_button = QtWidgets.QPushButton("Next >")
        self.cancel_button = QtWidgets.QPushButton("Cancel")
        self.help_button = QtWidgets.QPushButton("Help")
        self.back_button.clicked.connect(lambda: self.previous_tab(self.tab_widget))
        self.next_button.clicked.connect(lambda: self.next_tab(self.tab_widget))
        self.cancel_button.clicked.connect(QtWidgets.QApplication.quit)
        self.button_layout.addWidget(self.back_button)
        self.button_layout.addWidget(self.next_button)
        self.button_layout.addWidget(self.cancel_button)
        self.button_layout.addWidget(self.help_button)
        self.main_layout.addLayout(self.button_layout)
        self.date_to_files_map = defaultdict(list)
        self.headers = []
        self.folder_path = None
        self.client_name = None
        self.data_frames_map = {}
        self.uniqueness_results, self.uniqueness_exceptions = [], []
        self.omitted_records_summary, self.omitted_records_details = [], []

    def create_tabs(self):
        # Data Pre Processing
        data_pre_processing_tab = QtWidgets.QWidget()
        data_pre_processing_layout = QtWidgets.QVBoxLayout()
        data_pre_processing_label =QtWidgets.QLabel("""
<b>1. Standardized Data Structure:</b><br>
Ensure all files maintain a uniform structure to facilitate smooth compilation and analysis. Field headers must be in the first row and remain consistent across all datasets. For files containing multiple sheets, ensure the data tables in each sheet follow the same format. Any required data transformations or restructuring must be completed prior to using the application. Alteryx can be utilized for pre-processing and data clean-up steps to ensure consistency.<br>
<b>2. File Naming Convention:</b><br>
Rename files using the standardized format: <font color="red"><i>ClientName_AuditPhase_MM.DD.YYYY_SystemName</i></font> (e.g., RFM_YearEnd_03.31.2023_FISTDAccounts).This naming convention enables the application to accurately identify critical dataset attributes, such as the client name, system name, audit phase, and reporting date.<br>
<b>3. Input Folder Organization:</b><br>
Place files with identical structures into their respective input folders. Additionally, ensure all files within the same folder share the same extension type, either <i>.csv</i> or <i>.xlsx</i>, to avoid processing errors during analysis.<br>
<b>4. File Format Recommendations:</b><br>
For optimal performance during data validation, we recommend using <i>.csv</i> files instead of <i>.xlsx</i> files. The <i>.csv</i> format, being a simpler plain text structure without additional Excel formatting, enhances processing speed and ensures better compatibility with the application.
                                
                                                    
                                                    """)
        data_pre_processing_label.setWordWrap(True)
        data_pre_processing_layout.addWidget(data_pre_processing_label)                                                                  
        data_pre_processing_tab.setLayout(data_pre_processing_layout)
        self.tab_widget.addTab(data_pre_processing_tab, "Data Pre Processing")

        # Read Folder Tab
        read_folder_tab = QtWidgets.QWidget()
        read_folder_layout = QtWidgets.QVBoxLayout()
        read_folder_tab.setLayout(read_folder_layout)

        self.select_folder_btn = QtWidgets.QPushButton("Select Folder")
        self.select_folder_btn.clicked.connect(self.select_folder)
        read_folder_layout.addWidget(self.select_folder_btn)

        self.selected_folder_label = QtWidgets.QLabel("No folder selected")
        read_folder_layout.addWidget(self.selected_folder_label)

        # File List Section with QListWidget
        self.file_list_widget = QtWidgets.QListWidget()
        self.file_list_widget.itemClicked.connect(self.toggle_selection)
        read_folder_layout.addWidget(self.file_list_widget)

        # Select all and Deselect all buttons
        button_layout = QtWidgets.QHBoxLayout()
        self.select_all_btn = QtWidgets.QPushButton("Select All")
        self.select_all_btn.clicked.connect(lambda: self.select_all_files(self.file_list_widget))
        button_layout.addWidget(self.select_all_btn)

        self.deselect_all_btn = QtWidgets.QPushButton("Deselect All")
        self.deselect_all_btn.clicked.connect(lambda: self.deselect_all_files(self.file_list_widget))
        button_layout.addWidget(self.deselect_all_btn)

        read_folder_layout.addLayout(button_layout)

        self.tab_widget.addTab(read_folder_tab, "Read Folder")

        # Uniqueness Analysis Tab
        uniqueness_tab = QtWidgets.QWidget()
        uniqueness_layout = QtWidgets.QVBoxLayout()
        uniqueness_tab.setLayout(uniqueness_layout)

        choose_fields_layout = QtWidgets.QHBoxLayout()
        choose_unique_test_box = QtWidgets.QGroupBox("Choose Fields for Uniqueness Test")
        choose_unique_test_layout = QtWidgets.QVBoxLayout()
        self.header_list_widget = QtWidgets.QListWidget()
        self.header_list_widget.itemClicked.connect(self.toggle_selection)
        self.header_list_widget.itemClicked.connect(self.clear_tables)
        choose_unique_test_layout.addWidget(self.header_list_widget)
        choose_unique_test_box.setLayout(choose_unique_test_layout)
        choose_fields_layout.addWidget(choose_unique_test_box)

        choose_audit_box = QtWidgets.QGroupBox("Choose Field for Audit Balance")
        choose_audit_layout = QtWidgets.QVBoxLayout()
        self.audit_list_widget = QtWidgets.QListWidget()
        self.audit_list_widget.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        self.audit_list_widget.itemClicked.connect(lambda it: self.single_selection(self.audit_list_widget, it))
        self.audit_list_widget.itemClicked.connect(self.clear_tables)
        choose_audit_layout.addWidget(self.audit_list_widget)
        choose_audit_box.setLayout(choose_audit_layout)
        choose_fields_layout.addWidget(choose_audit_box)
        choose_fields_layout.setStretch(0, 1)  # 第一个控件的伸展系数
        choose_fields_layout.setStretch(1, 1)  # 第二个控件的伸展系数
        uniqueness_layout.addLayout(choose_fields_layout)

        self.uniqueness_tab_widget = QtWidgets.QTabWidget()
        self.uniqueness_field_summary_table_widget = QtWidgets.QTableWidget()
        self.uniqueness_field_summary_table_widget.resizeColumnsToContents()
        self.uniqueness_tab_widget.addTab(self.uniqueness_field_summary_table_widget, "Uniqueness Field Summary")

        self.uniqueness_exceptions_table_widget = QtWidgets.QTableWidget()
        self.uniqueness_exceptions_table_widget.resizeColumnsToContents()
        self.uniqueness_tab_widget.addTab(self.uniqueness_exceptions_table_widget, "Uniqueness Exceptions")
        uniqueness_layout.addWidget(self.uniqueness_tab_widget)

        # Button to perform uniqueness test
        uniqueness_btns_layout = QtWidgets.QHBoxLayout()
        self.uniqueness_test_btn = QtWidgets.QPushButton("Perform Uniqueness Test")
        self.uniqueness_test_btn.clicked.connect(self.perform_uniqueness_test)
        uniqueness_btns_layout.addWidget(self.uniqueness_test_btn)

        self.export_uniqueness_test_btn = QtWidgets.QPushButton("Export Uniqueness Test")
        self.export_uniqueness_test_btn.clicked.connect(self.export_uniqueness_test)

        uniqueness_btns_layout.addWidget(self.export_uniqueness_test_btn)
        self.top_exceptions_label = QLabel("Top Exceptions:")
        self.top_exceptions_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.top_exceptions_input = QSpinBox()
        self.top_exceptions_input.setRange(1, 20000)  # Set a reasonable range
        self.top_exceptions_input.setValue(1000)  # Default to 10
        uniqueness_btns_layout.addWidget(self.top_exceptions_label)
        uniqueness_btns_layout.addWidget(self.top_exceptions_input)
        uniqueness_layout.addLayout(uniqueness_btns_layout)

        self.tab_widget.addTab(uniqueness_tab, "Uniqueness Analysis")
        self.tab_widget.currentChanged.connect(self.on_tab_changed)
        # Rollforward Analysis Tab
        rollforward_tab = QtWidgets.QWidget()
        rollforward_layout = QtWidgets.QVBoxLayout()

# Canvas for Histogram
        self.rollforward_canvas = FigureCanvas(Figure(figsize=(5, 3)))
        rollforward_layout.addWidget(self.rollforward_canvas)

# Toolbar for Histogram
        self.rollforward_toolbar = NavigationToolbar(self.rollforward_canvas, self)
        rollforward_layout.addWidget(self.rollforward_toolbar)

# Perform Rollforward Analysis Button
        self.rollforward_button = QtWidgets.QPushButton("Perform Rollforward Analysis")
        self.rollforward_button.clicked.connect(self.perform_rollforward_analysis)
        rollforward_layout.addWidget(self.rollforward_button)

# Export Results Button
        self.export_rollforward_button = QtWidgets.QPushButton("Export Rollforward Results")
        self.export_rollforward_button.clicked.connect(
    lambda: self.export_rollforward_to_excel([])
)  # Placeholder, replace `[]` with stored results
        rollforward_layout.addWidget(self.export_rollforward_button)

        rollforward_tab.setLayout(rollforward_layout)
        self.tab_widget.addTab(rollforward_tab, "Rollforward Analysis")



        # Omitted Records Analysis Tab
        omitted_records_tab = QtWidgets.QWidget()
        omitted_records_layout = QtWidgets.QVBoxLayout()
        omitted_records_tab.setLayout(omitted_records_layout)

        self.omitted_records_tab_widget = QtWidgets.QTabWidget()
        self.omitted_records_summary_table_widget = QtWidgets.QTableWidget()
        self.omitted_records_summary_table_widget.resizeColumnsToContents()
        self.omitted_records_tab_widget.addTab(self.omitted_records_summary_table_widget, "Omitted Records Summary")

        self.omitted_records_details_table_widget = QtWidgets.QTableWidget()
        self.omitted_records_details_table_widget.resizeColumnsToContents()
        self.omitted_records_tab_widget.addTab(self.omitted_records_details_table_widget, "Omitted Records Details")
        omitted_records_layout.addWidget(self.omitted_records_tab_widget)

        # Button to perform omitted records test
        omitted_records_btns_layout = QtWidgets.QHBoxLayout()
        self.omitted_records_test_btn = QtWidgets.QPushButton("Perform Omitted Records Test")
        self.omitted_records_test_btn.clicked.connect(self.perform_omitted_test)
        omitted_records_btns_layout.addWidget(self.omitted_records_test_btn)

        # 增加一个导出数据的按钮
        self.export_omitted_records_test_btn = QtWidgets.QPushButton("Export Omitted Records Test")
        self.export_omitted_records_test_btn.clicked.connect(self.export_omitted_test)
        omitted_records_btns_layout.addWidget(self.export_omitted_records_test_btn)

        omitted_records_layout.addLayout(omitted_records_btns_layout)

        self.tab_widget.addTab(omitted_records_tab, "Omitted Records Analysis")

        # Cross-Period Analysis Tab
        cross_period_tab = QtWidgets.QWidget()
        cross_period_layout = QtWidgets.QVBoxLayout()
        cross_period_tab.setLayout(cross_period_layout)
        cross_period_label = QtWidgets.QLabel("Content for Cross-Period Analysis")
        # cross_period_layout.addWidget(cross_period_label)
        # self.tab_widget.addTab(cross_period_tab, "Cross-Period Analysis")

        # Profile Analysis Tab
        # Profile Analysis Tab
        profile_tab = QtWidgets.QWidget()
        profile_layout = QtWidgets.QVBoxLayout()

# Field and Datatype Selection Section
        field_selection_group = QtWidgets.QGroupBox("Select Fields and Datatypes")
        field_selection_layout = QtWidgets.QVBoxLayout()

# Field Selection Dropdown
        self.field_selection_label = QtWidgets.QLabel("Choose Fields:")
        self.field_selection_dropdown = QtWidgets.QComboBox()
        self.field_selection_dropdown.setPlaceholderText("Select a field")
        field_selection_layout.addWidget(self.field_selection_label)
        field_selection_layout.addWidget(self.field_selection_dropdown)

# Datatype Selection Dropdown
        self.datatype_selection_label = QtWidgets.QLabel("Choose Datatype:")
        self.datatype_selection_dropdown = QtWidgets.QComboBox()
        self.datatype_selection_dropdown.addItems(["Number", "Text", "Date"])
        field_selection_layout.addWidget(self.datatype_selection_label)
        field_selection_layout.addWidget(self.datatype_selection_dropdown)

        field_selection_group.setLayout(field_selection_layout)
        profile_layout.addWidget(field_selection_group)

# Button to perform profiling analysis
        self.profiling_button = QtWidgets.QPushButton("Perform Profiling Analysis")
        self.profiling_button.clicked.connect(self.perform_profiling_analysis)
        profile_layout.addWidget(self.profiling_button)

# Export Results Button
        self.export_profiling_button = QtWidgets.QPushButton("Export Profiling Results")
        self.export_profiling_button.clicked.connect(self.export_profiling_to_excel)
        profile_layout.addWidget(self.export_profiling_button)

# Results Display
        self.profiling_results = QtWidgets.QTextEdit()
        self.profiling_results.setReadOnly(True)
        profile_layout.addWidget(self.profiling_results)

        profile_tab.setLayout(profile_layout)
        self.tab_widget.addTab(profile_tab, "Profile Analysis")



    def clear_tables(self):
        table_widgets = [
            self.uniqueness_field_summary_table_widget,
            self.uniqueness_exceptions_table_widget,
            self.omitted_records_summary_table_widget,
            self.omitted_records_details_table_widget
        ]
        for tab in table_widgets:
            tab.clear()
            tab.setRowCount(0)
            tab.setColumnCount(0)

    def select_folder(self):
        folder = QtWidgets.QFileDialog.getExistingDirectory(self, "Select Folder")
        if folder:
            self.folder_path = folder
            self.selected_folder_label.setText(f"Selected Folder: {folder}")
            self.load_file_list()
        else:
            self.file_list_widget.clear()
            self.selected_folder_label.setText("No folder selected")

    @property
    def selected_dates(self):
        return [self.file_list_widget.item(i).text() for i in range(self.file_list_widget.count())
                if self.file_list_widget.item(i).checkState() == QtCore.Qt.Checked]

    @property
    def selected_headers(self):
        return [self.header_list_widget.item(i).text() for i in range(self.header_list_widget.count())
                if self.header_list_widget.item(i).checkState() == QtCore.Qt.Checked]

    @property
    def selected_audit_field(self):
        if self.audit_list_widget.currentItem():
            return self.audit_list_widget.currentItem().text()
        return None

    @property
    def top_exceptions(self):
        return self.top_exceptions_input.value() or 10

    def load_file_list(self):
        self.client_name, self.date_to_files_map = self.file_tools.load_files_by_folder(self.folder_path)
        self.file_list_widget.clear()
        for date_str in sorted(self.date_to_files_map.keys(), key=lambda x: datetime.strptime(x, '%m-%d-%Y')):
            if not date_str:
                continue
            item = QtWidgets.QListWidgetItem(date_str)
            item.setFlags(item.flags() | QtCore.Qt.ItemIsUserCheckable)
            item.setCheckState(QtCore.Qt.Unchecked)
            self.file_list_widget.addItem(item)

    def on_tab_changed(self, index):
     current_tab = self.tab_widget.tabText(index)

    # Update instructions
     self.instruction_label.setText(Constants.Instructions.get(current_tab, ""))
     self.instruction_label.setWordWrap(True)

    # Load fields dynamically for Profile Analysis
     if current_tab == "Profile Analysis":
        self.load_profile_fields()

     if self.tab_widget.tabText(index) == "Uniqueness Analysis":
        self.load_unique_headers_into_list()
        self.instruction_label.setText(Constants.Instructions.get(self.tab_widget.tabText(index), ""))
        self.instruction_label.setWordWrap(True)
        
    def load_profile_fields(self):
    # Clear the dropdown before loading new fields
     self.field_selection_dropdown.clear()

    # Ensure that selected dates are available
     selected_dates = self.selected_dates
     if not selected_dates:
        QtWidgets.QMessageBox.warning(self, "Warning", "No dates selected.")
        return

    # Load data and extract headers
     try:
        data_frames_map = self.file_tools.load_files_data_frames_map(self.date_to_files_map, selected_dates)
        headers = self.file_tools.load_data_frames_headers(data_frames_map.values())

        # Populate the dropdown with headers
        if headers:
            self.field_selection_dropdown.addItems(headers)
        else:
            QtWidgets.QMessageBox.warning(self, "Warning", "No fields available in the selected dataset.")
     except Exception as e:
        QtWidgets.QMessageBox.critical(self, "Error", f"Error loading fields: {str(e)}")


    def render_header_list_widget(self, headers):
        # 比较列表值是否相等
        if headers == self.headers:
            return
        self.header_list_widget.clear()
        self.audit_list_widget.clear()
        self.clear_tables()
        for header_str in headers:
            item = QtWidgets.QListWidgetItem(header_str)
            item.setFlags(item.flags() | QtCore.Qt.ItemIsUserCheckable)
            item.setCheckState(QtCore.Qt.Unchecked)
            self.header_list_widget.addItem(item)
            self.audit_list_widget.addItem(item.clone())
        self.headers = headers

    def load_unique_headers_into_list(self):
        selected_dates = self.selected_dates
        headers = []
        try:
            data_frames_map = self.file_tools.load_files_data_frames_map(self.date_to_files_map, selected_dates)
            headers = self.file_tools.load_data_frames_headers(data_frames_map.values())
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Error", f"Error reading file: {str(e)}")
        self.render_header_list_widget(headers)

    def build_uniqueness_summary_table(self, uniqueness_results):
        self.uniqueness_field_summary_table_widget.clear()
        self.uniqueness_field_summary_table_widget.setStyleSheet(
            "QHeaderView::section:horizontal { background-color: #FFE600; color: black; }")
        self.uniqueness_field_summary_table_widget.setRowCount(len(uniqueness_results))
        self.uniqueness_field_summary_table_widget.setColumnCount(len(Constants.UNIQUENESS_SUMMARY_HEADERS))
        self.uniqueness_field_summary_table_widget.setHorizontalHeaderLabels(Constants.UNIQUENESS_SUMMARY_HEADERS)

        for row, result in enumerate(uniqueness_results):
            for col, header in enumerate(Constants.UNIQUENESS_SUMMARY_HEADERS):
                col_str = str(result[col])
                item = QTableWidgetItem(col_str)
                if col in [2, 3]:
                    item.setTextAlignment(Qt.AlignCenter)
                    # 如果不唯一，则背景色标红
                    if col_str in ["Not Always Unique", "Not Unique"]:
                        item.setBackground(QBrush(QColor(255, 199, 206)))
                self.uniqueness_field_summary_table_widget.setItem(row, col, item)
        self.uniqueness_field_summary_table_widget.resizeColumnsToContents()

    def build_uniqueness_exceptions_table(self, uniqueness_exceptions):
        self.uniqueness_exceptions_table_widget.clear()
        self.uniqueness_exceptions_table_widget.setStyleSheet(
            "QHeaderView::section:horizontal { background-color: #FFE600; color: black; }")
        self.uniqueness_exceptions_table_widget.setRowCount(len(uniqueness_exceptions))
        self.uniqueness_exceptions_table_widget.setColumnCount(len(Constants.UNIQUENESS_EXCEPTIONS_HEADERS))
        self.uniqueness_exceptions_table_widget.setHorizontalHeaderLabels(Constants.UNIQUENESS_EXCEPTIONS_HEADERS)

        for row, exception in enumerate(uniqueness_exceptions):
            for col, header in enumerate(Constants.UNIQUENESS_EXCEPTIONS_HEADERS):
                try:
                    if col in [4, 5]:
                        exc_str = "${:>20,.2f}".format(exception[col])
                    elif col in [6]:
                        exc_str = "{:,.4%}".format(exception[col])
                    else:
                        exc_str = str(exception[col])
                except Exception as e:
                    exc_str = str(exception[col])
                item = QTableWidgetItem(exc_str)
                if col in [4, 5, 6]:
                    item.setTextAlignment(Qt.AlignRight|Qt.AlignVCenter)
                elif col in [3]:
                    item.setTextAlignment(Qt.AlignCenter)
                self.uniqueness_exceptions_table_widget.setItem(row, col, item)
        self.uniqueness_exceptions_table_widget.resizeColumnsToContents()

    def perform_uniqueness_test(self):
        selected_headers = self.selected_headers

        if not selected_headers:
            QtWidgets.QMessageBox.warning(self, "Warning", "No headers selected for uniqueness test.")
            return

        selected_dates = self.selected_dates
        self.data_frames_map = self.file_tools.load_files_data_frames_map(
            self.date_to_files_map,
            selected_dates,
            selected_headers + [self.selected_audit_field]
        )
        self.uniqueness_results, self.uniqueness_exceptions = self.math_tools.uniqueness_test(
            self.data_frames_map,
            selected_dates,
            selected_headers,
            self.selected_audit_field,
            self.top_exceptions
        )

        print(self.uniqueness_results)
        print(self.uniqueness_exceptions)
        self.build_uniqueness_summary_table(self.uniqueness_results)
        self.build_uniqueness_exceptions_table(self.uniqueness_exceptions)

    def export_uniqueness_test(self):
        self.export_to_excel(f"{self.client_name}_Uniqueness", {
            "Unique Field Summary": self.uniqueness_field_summary_table_widget,
            "Uniqueness Exceptions": self.uniqueness_exceptions_table_widget
        })

    def export_to_excel(self, name, tables):
        dlg = QFileDialog()
        options = dlg.options()
        output_filename = f"{name}_{datetime.now().strftime('%m%d%Y')}.xlsx"
        file_path, _ = dlg.getSaveFileName(
            None, "Save Summary",
            os.path.join(self.folder_path, output_filename) if self.folder_path else "",
            "Excel Files (*.xlsx)",
            options=options
        )
        if not file_path:
            return

        with pd.ExcelWriter(file_path) as writer:
            for sheet_name, table_widget in tables.items():
                headers = [table_widget.horizontalHeaderItem(i).text()
                           for i in range(table_widget.columnCount())]

                data = []
                for row in range(table_widget.rowCount()):
                    row_data = []
                    for col in range(table_widget.columnCount()):
                        item = table_widget.item(row, col)
                        row_data.append(item.text() if item else "")
                    data.append(row_data)

                df = pd.DataFrame(data, columns=headers)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                worksheet = writer.sheets[sheet_name]
                fill = PatternFill(start_color="FFE600", fill_type="solid")
                for col_idx, col in enumerate(df.columns, start=1):
                    cell = worksheet.cell(row=1, column=col_idx)  # 表头位于第一行
                    cell.fill = fill
                for col_idx, col in enumerate(df.columns, start=1):
                    max_length = max(
                        df[col].astype(str).map(len).max(),
                        len(str(col))
                    )
                    worksheet.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

    def build_omitted_summary_table(self, omitted_records_summary):
        self.omitted_records_summary_table_widget.clear()
        self.omitted_records_summary_table_widget.setStyleSheet(
            "QHeaderView::section:horizontal { background-color: #FFE600; color: black; }")
        self.omitted_records_summary_table_widget.setRowCount(len(omitted_records_summary))
        self.omitted_records_summary_table_widget.setColumnCount(len(Constants.OMITTED_RECORDS_SUMMARY_HEADERS))
        self.omitted_records_summary_table_widget.setHorizontalHeaderLabels(Constants.OMITTED_RECORDS_SUMMARY_HEADERS)
        for row, (key, value) in enumerate(omitted_records_summary.items()):
            for col, header in enumerate(Constants.OMITTED_RECORDS_SUMMARY_HEADERS):
                item = QTableWidgetItem(str(value) if col == 1 else key)
                if col in [1]:
                    item.setTextAlignment(Qt.AlignCenter)
                self.omitted_records_summary_table_widget.setItem(row, col, item)
        self.omitted_records_summary_table_widget.resizeColumnsToContents()

    def build_omitted_details_table(self, omitted_records_details):
        self.omitted_records_details_table_widget.clear()
        self.omitted_records_details_table_widget.setStyleSheet(
            "QHeaderView::section:horizontal { background-color: #FFE600; color: black; }")
        self.omitted_records_details_table_widget.setRowCount(len(omitted_records_details))
        self.omitted_records_details_table_widget.setColumnCount(len(Constants.OMITTED_RECORDS_DETAILS_HEADERS))
        self.omitted_records_details_table_widget.setHorizontalHeaderLabels(Constants.OMITTED_RECORDS_DETAILS_HEADERS)
        for row, details in enumerate(omitted_records_details):
            for col, header in enumerate(Constants.OMITTED_RECORDS_DETAILS_HEADERS):
                item = QTableWidgetItem(str(details[col]))
                if col in [1]:
                    item.setTextAlignment(Qt.AlignCenter)
                self.omitted_records_details_table_widget.setItem(row, col, item)
        self.omitted_records_details_table_widget.resizeColumnsToContents()

    def perform_omitted_test(self):
        selected_headers = self.selected_headers

        if not selected_headers:
            QtWidgets.QMessageBox.warning(self, "Warning", "No headers selected for omitted test.")
            return

        selected_dates = self.selected_dates
        self.data_frames_map = self.file_tools.load_files_data_frames_map(
            self.date_to_files_map,
            selected_dates,
            selected_headers
        )
        print(self.data_frames_map)
        self.omitted_records_summary, self.omitted_records_details = self.math_tools.omitted_records_test(
            self.data_frames_map,
            selected_dates,
            selected_headers
        )

        print(self.omitted_records_summary)
        print(self.omitted_records_details)
        self.build_omitted_details_table(self.omitted_records_details)
        self.build_omitted_summary_table(self.omitted_records_summary)

    def export_omitted_test(self):
        self.export_to_excel(f"{self.client_name}_Omitted", {
            "Omitted Records Summary": self.omitted_records_summary_table_widget,
            "Omitted Records Details": self.omitted_records_details_table_widget
        })

    def perform_rollforward_analysis(self):
     selected_dates = self.selected_dates
     unique_id_field = self.selected_headers[0] if self.selected_headers else None
     audit_field = self.selected_audit_field

     if not selected_dates or not unique_id_field or not audit_field:
        QtWidgets.QMessageBox.warning(
            self, "Warning", "Please select dates, a unique ID field, and an audit field."
        )
        return

     try:
        data_frames_map = self.file_tools.load_files_data_frames_map(
            self.date_to_files_map, selected_dates
        )
        results = self.math_tools.rollforward_analysis(
            data_frames_map, selected_dates, unique_id_field, audit_field
        )

        # Prepare data for histogram
        added_counts = [result["Added Count"] for result in results]
        dropped_counts = [result["Dropped Count"] for result in results]
        labels = [f"{result['Previous Date']} to {result['Current Date']}" for result in results]

        # Plot the histogram
        self.plot_rollforward_histogram(labels, added_counts, dropped_counts)

        # Export results if needed
        self.export_rollforward_to_excel(results)
        QtWidgets.QMessageBox.information(self, "Success", "Rollforward analysis completed successfully.")
     except Exception as e:
        QtWidgets.QMessageBox.critical(self, "Error", f"Error during rollforward analysis: {str(e)}")

    def plot_rollforward_histogram(self, labels, added_counts, dropped_counts):
    # Clear the canvas
     self.rollforward_canvas.figure.clear()

    # Create subplots
     ax = self.rollforward_canvas.figure.add_subplot(111)

    # Bar widths and positions
     x = range(len(labels))
     bar_width = 0.4

    # Plot Added and Dropped Counts
     ax.bar(x, added_counts, bar_width, label="Added", color="green")
     ax.bar([i + bar_width for i in x], dropped_counts, bar_width, label="Dropped", color="red")

    # Add labels and title
     ax.set_xticks([i + bar_width / 2 for i in x])
     ax.set_xticklabels(labels, rotation=45, ha="right")
     ax.set_xlabel("Date Ranges")
     ax.set_ylabel("Counts")
     ax.set_title("Rollforward Analysis")
     ax.legend()

    # Redraw the canvas
     self.rollforward_canvas.draw()

    # def perform_rollforward_analysis(self):
    #  selected_dates = self.selected_dates
    #  unique_id_field = self.selected_headers[0] if self.selected_headers else None
    #  audit_field = self.selected_audit_field

    #  if not selected_dates or not unique_id_field or not audit_field:
    #     QtWidgets.QMessageBox.warning(self, "Warning", "Please select dates, a unique ID field, and an audit field.")
    #     return

    #  try:
    #     data_frames_map = self.file_tools.load_files_data_frames_map(self.date_to_files_map, selected_dates)
    #     # Ensure only 4 arguments are passed
    #     # print(selected_dates)
    #     # print(unique_id_field)
    #     # print(audit_field)
    #     print(data_frames_map)
    #     results = self.math_tools.rollforward_analysis(data_frames_map, selected_dates, unique_id_field, audit_field)
    #     # print(results+'hello')
    #     # Display results
    #     result_text = "\n".join([f"{key}: {value}" for summary in results for key, value in summary.items()])
    #     self.rollforward_results.setText(result_text)

    #     # Generate report and export results
    #     # self.generate_rollforward_report(results)
    #     self.export_rollforward_to_excel(results)

    #     QtWidgets.QMessageBox.information(self, "Success", "Rollforward analysis completed successfully.")
    #  except Exception as e:
    #     QtWidgets.QMessageBox.critical(self, "Error", f"Error during rollforward analysis: {str(e)}")


    def perform_profiling_analysis(self):
     selected_dates = self.selected_dates
     selected_fields = self.selected_headers

     if not selected_dates or not selected_fields:
        QtWidgets.QMessageBox.warning(self, "Warning", "Please select dates and fields for profiling analysis.")
        return

     try:
        data_frames_map = self.file_tools.load_files_data_frames_map(self.date_to_files_map, selected_dates)
        profiling_results = self.math_tools.profiling_analysis(data_frames_map, selected_dates, selected_fields)

        # Display results
        result_text = ""
        for date, results in profiling_results.items():
            result_text += f"\nDate: {date}\n"
            for field_result in results:
                result_text += "\n".join([f"{key}: {value}" for key, value in field_result.items()])
                result_text += "\n\n"
        self.profiling_results.setText(result_text)

        # Save profiling results for export
        self.profiling_results_data = profiling_results

        QtWidgets.QMessageBox.information(self, "Success", "Profiling analysis completed.")
     except Exception as e:
        QtWidgets.QMessageBox.critical(self, "Error", f"Error during profiling analysis: {str(e)}")

    def export_profiling_to_excel(self):
    # import pandas as pd
    # from PySide6.QtWidgets import QFileDialog

     if not hasattr(self, 'profiling_results_data') or not self.profiling_results_data:
        QtWidgets.QMessageBox.warning(self, "Warning", "No profiling results to export.")
        return

    # Select file save location
     options = QFileDialog.Options()
     file_path, _ = QFileDialog.getSaveFileName(
        self,
        "Save Profiling Results",
        f"Profiling_Analysis_{datetime.now().strftime('%Y%m%d')}.xlsx",
        "Excel Files (*.xlsx)",
        options=options,
    )

     if not file_path:
        return

    # Convert results to Excel
     try:
        with pd.ExcelWriter(file_path) as writer:
            for date, results in self.profiling_results_data.items():
                df = pd.DataFrame(results)
                df.to_excel(writer, index=False, sheet_name=date)

        QtWidgets.QMessageBox.information(self, "Export Successful", f"Profiling results saved to {file_path}")
     except Exception as e:
        QtWidgets.QMessageBox.critical(self, "Export Failed", f"Failed to export profiling results: {str(e)}")

    def export_rollforward_to_excel(self, results):
    # import pandas as pd
    # from PySide6.QtWidgets import QFileDialog

    # Convert results to a DataFrame
     df = pd.DataFrame(results)

    # Prompt user for file save location
     options = QFileDialog.Options()
     file_path, _ = QFileDialog.getSaveFileName(
        self,
        "Save Rollforward Results",
        f"Rollforward_Analysis_{datetime.now().strftime('%Y%m%d')}.xlsx",
        "Excel Files (*.xlsx)",
        options=options,
    )

     if not file_path:
        return

    # Save DataFrame to Excel
     try:
        with pd.ExcelWriter(file_path) as writer:
            df.to_excel(writer, index=False, sheet_name="Rollforward Analysis")
        QtWidgets.QMessageBox.information(self, "Export Successful", f"Rollforward results saved to {file_path}")
     except Exception as e:
        QtWidgets.QMessageBox.critical(self, "Export Failed", f"Failed to export results: {str(e)}")


if __name__ == "__main__":
    import sys

    app = QtWidgets.QApplication.instance()
    if app is None:
        app = QtWidgets.QApplication(sys.argv)
    window = FileConsolidationApp()
    window.show()
    sys.exit(app.exec())
